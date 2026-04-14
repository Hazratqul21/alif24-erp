import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
CELL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_to_excel(headers: list[str], rows: list[list], sheet_name: str = "Ma'lumotlar") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = CELL_BORDER

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
