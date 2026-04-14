import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.core.dependencies import get_db, get_current_user
from app.core.auth import hash_password
from app.core.rbac import check_permission
from app.core.exceptions import (
    NotFoundError,
    ForbiddenError,
    ConflictError,
    ValidationError,
)
from app.schemas.user import (
    UserCreate,
    UserUpdate,
    UserResponse,
    UserListResponse,
)

router = APIRouter(tags=["Users"])


def _build_user_response(row, roles: list[str]) -> UserResponse:
    return UserResponse(
        id=str(row[0]),
        email=row[1],
        phone=row[2],
        first_name=row[3],
        last_name=row[4],
        middle_name=row[5],
        birth_date=row[6],
        gender=row[7],
        avatar=row[8],
        is_active=row[9],
        roles=roles,
        created_at=row[10],
    )


async def _get_user_roles(db: AsyncSession, user_id: str) -> list[str]:
    result = await db.execute(
        text("""
            SELECT r.name FROM roles r
            JOIN user_roles ur ON ur.role_id = r.id
            WHERE ur.user_id = :uid
        """),
        {"uid": user_id},
    )
    return [row[0] for row in result]


@router.get("/", response_model=UserListResponse)
async def list_users(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: str = Query(None, description="Ism, email yoki telefon bo'yicha qidirish"),
    role: str = Query(None, description="Rol nomi bo'yicha filtrlash"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    if not check_permission(current_user, "users", "view"):
        raise ForbiddenError("'users.view' ruxsati yo'q")

    conditions = ["u.deleted_at IS NULL"]
    params: dict = {}

    if search:
        conditions.append("""
            (u.first_name ILIKE :search
             OR u.last_name ILIKE :search
             OR u.email ILIKE :search
             OR u.phone ILIKE :search)
        """)
        params["search"] = f"%{search}%"

    if role:
        conditions.append("""
            EXISTS (
                SELECT 1 FROM user_roles ur2
                JOIN roles r2 ON r2.id = ur2.role_id
                WHERE ur2.user_id = u.id AND r2.name = :role_name
            )
        """)
        params["role_name"] = role

    where_clause = " AND ".join(conditions)

    count_result = await db.execute(
        text(f"SELECT COUNT(*) FROM users u WHERE {where_clause}"),
        params,
    )
    total = count_result.scalar() or 0

    offset = (page - 1) * per_page
    params["limit"] = per_page
    params["offset"] = offset

    result = await db.execute(
        text(f"""
            SELECT u.id, u.email, u.phone, u.first_name, u.last_name,
                   u.middle_name, u.birth_date, u.gender, u.avatar,
                   u.is_active, u.created_at
            FROM users u
            WHERE {where_clause}
            ORDER BY u.created_at DESC
            LIMIT :limit OFFSET :offset
        """),
        params,
    )
    rows = result.fetchall()

    items = []
    for row in rows:
        roles = await _get_user_roles(db, str(row[0]))
        items.append(_build_user_response(row, roles))

    total_pages = (total + per_page - 1) // per_page if total > 0 else 0

    return UserListResponse(
        items=items,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )


@router.post("/", response_model=UserResponse, status_code=201)
async def create_user(
    body: UserCreate,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    if not check_permission(current_user, "users", "create"):
        raise ForbiddenError("'users.create' ruxsati yo'q")

    if body.email:
        dup = await db.execute(
            text("SELECT id FROM users WHERE email = :email AND deleted_at IS NULL"),
            {"email": body.email},
        )
        if dup.fetchone():
            raise ConflictError("Bu email allaqachon ro'yxatdan o'tgan")

    if body.phone:
        dup = await db.execute(
            text("SELECT id FROM users WHERE phone = :phone AND deleted_at IS NULL"),
            {"phone": body.phone},
        )
        if dup.fetchone():
            raise ConflictError("Bu telefon raqam allaqachon ro'yxatdan o'tgan")

    password_hash = hash_password(body.password)

    result = await db.execute(
        text("""
            INSERT INTO users (id, email, phone, password_hash, first_name, last_name,
                               middle_name, birth_date, gender, is_active, created_at, updated_at)
            VALUES (gen_random_uuid()::text, :email, :phone, :ph, :fn, :ln,
                    :mn, :bd, :gender, true, NOW(), NOW())
            RETURNING id, email, phone, first_name, last_name, middle_name,
                      birth_date, gender, avatar, is_active, created_at
        """),
        {
            "email": body.email,
            "phone": body.phone,
            "ph": password_hash,
            "fn": body.first_name,
            "ln": body.last_name,
            "mn": body.middle_name,
            "bd": body.birth_date,
            "gender": body.gender,
        },
    )
    new_user = result.fetchone()

    user_id = str(new_user[0])
    roles: list[str] = []

    if body.role_ids:
        for role_id in body.role_ids:
            role_check = await db.execute(
                text("SELECT name FROM roles WHERE id = :rid"),
                {"rid": role_id},
            )
            role_row = role_check.fetchone()
            if role_row:
                await db.execute(
                    text("""
                        INSERT INTO user_roles (user_id, role_id, assigned_by, assigned_at)
                        VALUES (:uid, :rid, :assigned_by, NOW())
                        ON CONFLICT DO NOTHING
                    """),
                    {"uid": user_id, "rid": role_id, "assigned_by": current_user["id"]},
                )
                roles.append(role_row[0])

    await db.commit()

    return _build_user_response(new_user, roles)


@router.get("/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    if not check_permission(current_user, "users", "view"):
        raise ForbiddenError("'users.view' ruxsati yo'q")

    result = await db.execute(
        text("""
            SELECT id, email, phone, first_name, last_name, middle_name,
                   birth_date, gender, avatar, is_active, created_at
            FROM users
            WHERE id = :uid AND deleted_at IS NULL
        """),
        {"uid": user_id},
    )
    row = result.fetchone()
    if not row:
        raise NotFoundError("Foydalanuvchi")

    roles = await _get_user_roles(db, user_id)
    return _build_user_response(row, roles)


@router.put("/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: str,
    body: UserUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    if not check_permission(current_user, "users", "edit"):
        raise ForbiddenError("'users.edit' ruxsati yo'q")

    check = await db.execute(
        text("SELECT id FROM users WHERE id = :uid AND deleted_at IS NULL"),
        {"uid": user_id},
    )
    if not check.fetchone():
        raise NotFoundError("Foydalanuvchi")

    updates = body.model_dump(exclude_unset=True, exclude={"role_ids"})

    if "email" in updates and updates["email"]:
        dup = await db.execute(
            text("SELECT id FROM users WHERE email = :email AND id != :uid AND deleted_at IS NULL"),
            {"email": updates["email"], "uid": user_id},
        )
        if dup.fetchone():
            raise ConflictError("Bu email allaqachon boshqa foydalanuvchida mavjud")

    if "phone" in updates and updates["phone"]:
        dup = await db.execute(
            text("SELECT id FROM users WHERE phone = :phone AND id != :uid AND deleted_at IS NULL"),
            {"phone": updates["phone"], "uid": user_id},
        )
        if dup.fetchone():
            raise ConflictError("Bu telefon raqam allaqachon boshqa foydalanuvchida mavjud")

    if updates:
        set_parts = []
        params = {"uid": user_id}
        for key, val in updates.items():
            set_parts.append(f"{key} = :{key}")
            params[key] = val
        set_parts.append("updated_at = NOW()")
        set_clause = ", ".join(set_parts)

        await db.execute(
            text(f"UPDATE users SET {set_clause} WHERE id = :uid"),
            params,
        )

    if body.role_ids is not None:
        await db.execute(
            text("DELETE FROM user_roles WHERE user_id = :uid"),
            {"uid": user_id},
        )
        for role_id in body.role_ids:
            await db.execute(
                text("""
                    INSERT INTO user_roles (user_id, role_id, assigned_by, assigned_at)
                    VALUES (:uid, :rid, :assigned_by, NOW())
                    ON CONFLICT DO NOTHING
                """),
                {"uid": user_id, "rid": role_id, "assigned_by": current_user["id"]},
            )

    await db.commit()

    result = await db.execute(
        text("""
            SELECT id, email, phone, first_name, last_name, middle_name,
                   birth_date, gender, avatar, is_active, created_at
            FROM users
            WHERE id = :uid
        """),
        {"uid": user_id},
    )
    row = result.fetchone()
    roles = await _get_user_roles(db, user_id)
    return _build_user_response(row, roles)


@router.delete("/{user_id}")
async def delete_user(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    if not check_permission(current_user, "users", "delete"):
        raise ForbiddenError("'users.delete' ruxsati yo'q")

    if user_id == current_user["id"]:
        raise ValidationError("O'zingizni o'chira olmaysiz")

    result = await db.execute(
        text("SELECT id FROM users WHERE id = :uid AND deleted_at IS NULL"),
        {"uid": user_id},
    )
    if not result.fetchone():
        raise NotFoundError("Foydalanuvchi")

    await db.execute(
        text("""
            UPDATE users
            SET deleted_at = :now, is_active = false, updated_at = NOW()
            WHERE id = :uid
        """),
        {"now": datetime.now(timezone.utc), "uid": user_id},
    )
    await db.commit()

    return {"detail": "Foydalanuvchi muvaffaqiyatli o'chirildi"}


@router.post("/import")
async def import_users(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    if not check_permission(current_user, "users", "create"):
        raise ForbiddenError("'users.create' ruxsati yo'q")

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise ValidationError("Faqat Excel (.xlsx, .xls) fayl yuklash mumkin")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValidationError("Excel kutubxonasi o'rnatilmagan")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    rows_list = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows_list:
        raise ValidationError("Fayl bo'sh yoki ma'lumotlar topilmadi")

    created_count = 0
    errors = []

    for idx, row in enumerate(rows_list, start=2):
        if len(row) < 4:
            errors.append(f"{idx}-qator: kamida 4 ta ustun kerak (ism, familiya, telefon, parol)")
            continue

        first_name = str(row[0] or "").strip()
        last_name = str(row[1] or "").strip()
        phone = str(row[2] or "").strip()
        password = str(row[3] or "").strip()
        email = str(row[4] or "").strip() if len(row) > 4 and row[4] else None

        if not first_name or not last_name or not phone or not password:
            errors.append(f"{idx}-qator: ism, familiya, telefon va parol majburiy")
            continue

        dup = await db.execute(
            text("SELECT id FROM users WHERE phone = :phone AND deleted_at IS NULL"),
            {"phone": phone},
        )
        if dup.fetchone():
            errors.append(f"{idx}-qator: {phone} telefon raqam allaqachon mavjud")
            continue

        password_hash = hash_password(password)

        await db.execute(
            text("""
                INSERT INTO users (id, phone, email, password_hash, first_name, last_name,
                                   is_active, created_at, updated_at)
                VALUES (gen_random_uuid()::text, :phone, :email, :ph, :fn, :ln,
                        true, NOW(), NOW())
            """),
            {
                "phone": phone,
                "email": email,
                "ph": password_hash,
                "fn": first_name,
                "ln": last_name,
            },
        )
        created_count += 1

    await db.commit()

    return {
        "detail": f"{created_count} ta foydalanuvchi muvaffaqiyatli import qilindi",
        "total_rows": len(rows_list),
        "created": created_count,
        "errors": errors,
    }
